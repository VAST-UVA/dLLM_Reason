"""Benchmark evaluators for MBPP, HumanEval, HotpotQA, MMLU, GSM8K, MATH, ARC, ProntoQA.

Each evaluator:
1. Loads the dataset
2. Formats prompts for LLaDA
3. Generates responses using a given scheduler
4. Computes the benchmark metric
5. Returns a detailed results dict
6. Optionally saves per-sample outputs to JSON and/or Excel (--save_outputs)
"""

from __future__ import annotations

import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

import torch
from tqdm import tqdm

from dllm_reason.eval.metrics import (
    exact_match, f1_score, extract_number,
    extract_multiple_choice, pass_at_k,
    normalize_answer,
)
from dllm_reason.utils.logging import get_logger

logger = get_logger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# Output saver
# ──────────────────────────────────────────────────────────────────────────────

def _save_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    logger.info(f"Saved JSON: {path}")


def _save_xlsx(path: Path, rows: list[dict], columns: list[str]) -> None:
    """Write a list of row dicts to an Excel file with openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed — skipping Excel save. pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=False)

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            val = row.get(col_name, "")
            # Truncate very long strings so Excel stays usable
            if isinstance(val, str) and len(val) > 2000:
                val = val[:2000] + "…"
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    logger.info(f"Saved Excel: {path}")


# ──────────────────────────────────────────────────────────────────────────────
# Base class
# ──────────────────────────────────────────────────────────────────────────────

class BenchmarkEvaluator:
    """Abstract benchmark evaluator."""

    def __init__(
        self,
        model,
        scheduler,
        # ── inference ──────────────────────────────────────────────────────────
        num_steps: int = 128,
        block_length: int = 32,
        temperature: float = 0.0,
        cfg_scale: float = 0.0,
        remasking: str = "low_confidence",
        batch_size: int = 1,
        max_new_tokens: int = 128,
        num_samples: int | None = None,
        run_tests: bool = True,
        verbose_errors: bool = False,
        # ── output saving ──────────────────────────────────────────────────────
        save_outputs: bool = False,       # master switch
        save_dir: str | Path | None = None,
        save_qa: bool = True,             # include prompt + generated in output
        save_ground_truth: bool = True,   # include reference answers in output
        record_trajectory: bool = False,  # record per-step unmasking states
        output_formats: list[str] | None = None,  # ["json", "xlsx"]
        run_tag: str = "",                # tag appended to output filenames
    ):
        self.model = model
        self.scheduler = scheduler
        self.num_steps = num_steps
        self.block_length = block_length
        self.temperature = temperature
        self.cfg_scale = cfg_scale
        self.remasking = remasking
        self.batch_size = batch_size
        self.max_new_tokens = max_new_tokens
        self.num_samples = num_samples
        self.run_tests = run_tests
        self.verbose_errors = verbose_errors

        # saving
        self.save_outputs = save_outputs
        self.save_dir = Path(save_dir) if save_dir else Path("results")
        self.save_qa = save_qa
        self.save_ground_truth = save_ground_truth
        self.record_trajectory = record_trajectory
        self.output_formats = output_formats or ["json", "xlsx"]
        self.run_tag = run_tag

    # ── generation ─────────────────────────────────────────────────────────────

    def _generate(
        self, prompt: str, system_prompt: str | None = None
    ) -> tuple[str, list[str] | None]:
        """Run generation. Returns (text, trajectory_or_None)."""
        result = self.model.generate(
            prompt,
            generation_len=self.max_new_tokens,
            block_length=self.block_length,
            scheduler=self.scheduler,
            num_steps=self.num_steps,
            temperature=self.temperature,
            cfg_scale=self.cfg_scale,
            remasking=self.remasking,
            system_prompt=system_prompt,
            record_trajectory=self.record_trajectory,
        )
        if self.record_trajectory:
            # model returns (text, list[str])
            return result[0], result[1]
        return result, None

    # ── saving ──────────────────────────────────────────────────────────────────

    def _output_stem(self, benchmark: str) -> str:
        """Base filename (no extension) for this run's output files."""
        parts = [benchmark]
        if self.run_tag:
            parts.append(self.run_tag)
        return "_".join(parts)

    def _save_results(
        self,
        samples: list[dict],
        benchmark: str,
        xlsx_columns: list[str],
        summary: dict,
    ) -> None:
        """Write per-sample outputs to the configured formats."""
        if not self.save_outputs:
            return

        stem = self._output_stem(benchmark)

        if "json" in self.output_formats:
            _save_json(
                self.save_dir / f"{stem}_samples.json",
                {"summary": summary, "samples": samples},
            )

        if "xlsx" in self.output_formats:
            _save_xlsx(
                self.save_dir / f"{stem}_samples.xlsx",
                samples,
                xlsx_columns,
            )

        # Trajectory is large — always saved to its own file
        if self.record_trajectory and "json" in self.output_formats:
            traj_data = [
                {"idx": s.get("task_id", i), "trajectory": s.get("trajectory")}
                for i, s in enumerate(samples)
                if s.get("trajectory") is not None
            ]
            if traj_data:
                _save_json(
                    self.save_dir / f"{stem}_trajectory.json",
                    traj_data,
                )

    def evaluate(self) -> dict[str, Any]:
        raise NotImplementedError


# ──────────────────────────────────────────────────────────────────────────────
# MBPP
# ──────────────────────────────────────────────────────────────────────────────

class MBPPEvaluator(BenchmarkEvaluator):
    """MBPP: Mostly Basic Programming Problems.

    Metrics: pass@1
    Format: Given a docstring + test cases, generate a Python function.
    """

    SYSTEM_PROMPT = (
        "You are an expert Python programmer. "
        "Write a complete Python function based on the given description. "
        "Output only the function code, no explanation."
    )

    # Excel columns — order determines column order in the sheet
    XLSX_COLUMNS = [
        "task_id", "prompt", "ground_truth_code",
        "generated_code", "passed", "error", "stdout",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        dataset = resolve_dataset("google-research-datasets/mbpp", config="sanitized", split="test")
        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for item in tqdm(items, desc="MBPP"):
            task_id = item.get("task_id", item.get("source_file", str(len(results))))
            prompt = item.get("prompt", item.get("text", ""))
            test_list = item["test_list"]
            canonical = item["code"]

            # Include test cases so the model knows the function name/signature
            prompt_with_tests = (
                f"{prompt}\n\n"
                f"Your function must pass these tests:\n"
                + "\n".join(f"    {t}" for t in test_list)
            )

            generated, trajectory = self._generate(prompt_with_tests, self.SYSTEM_PROMPT)
            code = self._extract_code(generated)

            if len(results) < 3:
                logger.debug(
                    f"MBPP #{task_id} raw output ({len(generated)} chars):\n"
                    f"{generated[:500]}\n---extracted---\n{code[:500]}"
                )

            run_info = (
                self._run_tests(code, test_list)
                if self.run_tests
                else {"passed": False, "timed_out": False,
                      "stderr": "", "stdout": "", "error": ""}
            )

            if not run_info["passed"] and self.verbose_errors:
                logger.warning(
                    f"[MBPP {task_id}] FAILED\n"
                    f"  extracted_code: {code[:300]!r}\n"
                    f"  raw_output:     {generated[:200]!r}"
                )
                if run_info["timed_out"]:
                    logger.warning(f"[MBPP {task_id}] TIMEOUT")
                elif run_info["stderr"]:
                    logger.warning(f"[MBPP {task_id}] STDERR: {run_info['stderr'][:300]}")
                elif run_info["error"]:
                    logger.warning(f"[MBPP {task_id}] ERROR: {run_info['error']}")

            sample: dict[str, Any] = {
                "task_id": task_id,
                "passed": run_info["passed"],
                "pass@1": pass_at_k(1, int(run_info["passed"]), 1),
                "timed_out": run_info["timed_out"],
                "error": run_info["error"],
                "stdout": run_info["stdout"],
                "stderr": run_info["stderr"],
            }
            if self.save_qa:
                sample["prompt"] = prompt_with_tests
                sample["generated_code"] = code
                sample["raw_output"] = generated
            if self.save_ground_truth:
                sample["ground_truth_code"] = canonical
                sample["test_list"] = test_list
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            # Row for Excel (always human-readable subset)
            xlsx_rows.append({
                "task_id": str(task_id),
                "prompt": prompt_with_tests,
                "ground_truth_code": canonical,
                "generated_code": code,
                "passed": str(run_info["passed"]),
                "error": (run_info["stderr"] or run_info["error"] or "")[:500],
                "stdout": run_info["stdout"][:300],
            })

        pass_1 = sum(r["pass@1"] for r in results) / len(results)
        logger.info(f"MBPP pass@1: {pass_1:.4f} ({len(results)} problems)")

        summary = {
            "benchmark": "mbpp",
            "pass@1": pass_1,
            "num_problems": len(results),
        }
        self._save_results(xlsx_rows, "mbpp", self.XLSX_COLUMNS, summary)

        return {**summary, "per_problem": results}

    def _extract_code(self, text: str) -> str:
        """Extract Python code from model output."""
        code_block = re.search(r"```python\n(.*?)```", text, re.DOTALL)
        if code_block:
            return code_block.group(1)
        code_block = re.search(r"```\n(.*?)```", text, re.DOTALL)
        if code_block:
            return code_block.group(1)
        return text

    def _run_tests(self, code: str, tests: list[str], timeout: int = 10) -> dict:
        """Execute code + tests in a subprocess."""
        test_code = "\n".join(tests)
        full_code = f"{code}\n\n{test_code}\n"
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
                f.write(full_code)
                tmp_path = f.name
            result = subprocess.run(
                [sys.executable, tmp_path],
                capture_output=True, text=True, timeout=timeout,
            )
            return {
                "passed": result.returncode == 0,
                "timed_out": False,
                "stderr": result.stderr.strip(),
                "stdout": result.stdout.strip(),
                "error": "",
            }
        except subprocess.TimeoutExpired:
            return {"passed": False, "timed_out": True,
                    "stderr": "", "stdout": "", "error": f"timeout>{timeout}s"}
        except Exception as exc:
            return {"passed": False, "timed_out": False,
                    "stderr": "", "stdout": "", "error": str(exc)}
        finally:
            if tmp_path:
                Path(tmp_path).unlink(missing_ok=True)


# ──────────────────────────────────────────────────────────────────────────────
# HumanEval
# ──────────────────────────────────────────────────────────────────────────────

class HumanEvalEvaluator(BenchmarkEvaluator):
    """HumanEval: Python coding benchmark.

    Metrics: pass@1
    Format: Complete a function given its signature + docstring.
    """

    SYSTEM_PROMPT = (
        "Complete the following Python function. "
        "Output only the complete function implementation, no explanation."
    )

    XLSX_COLUMNS = [
        "task_id", "prompt", "canonical_solution",
        "generated_completion", "passed", "error", "stdout",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        dataset = resolve_dataset("openai/openai_humaneval", split="test")
        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for item in tqdm(items, desc="HumanEval"):
            task_id = item["task_id"]
            prompt = item["prompt"]
            canonical_solution = item["canonical_solution"]
            test = item["test"]
            entry_point = item["entry_point"]

            generated, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
            completion = self._extract_completion(generated, prompt)
            full_code = prompt + completion

            if len(results) < 3:
                logger.debug(
                    f"HumanEval {task_id} raw output ({len(generated)} chars):\n"
                    f"{generated[:500]}\n---completion---\n{completion[:500]}"
                )

            run_info = (
                self._run_humaneval_test(full_code, test, entry_point)
                if self.run_tests
                else {"passed": False, "timed_out": False,
                      "stderr": "", "stdout": "", "error": ""}
            )

            if not run_info["passed"] and self.verbose_errors:
                logger.warning(
                    f"[HumanEval {task_id}] FAILED\n"
                    f"  extracted_completion: {completion[:300]!r}\n"
                    f"  raw_output:           {generated[:200]!r}"
                )
                if run_info["timed_out"]:
                    logger.warning(f"[HumanEval {task_id}] TIMEOUT")
                elif run_info["stderr"]:
                    logger.warning(f"[HumanEval {task_id}] STDERR: {run_info['stderr'][:300]}")
                elif run_info["error"]:
                    logger.warning(f"[HumanEval {task_id}] ERROR: {run_info['error']}")

            sample: dict[str, Any] = {
                "task_id": task_id,
                "passed": run_info["passed"],
                "pass@1": pass_at_k(1, int(run_info["passed"]), 1),
                "timed_out": run_info["timed_out"],
                "error": run_info["error"],
                "stdout": run_info["stdout"],
                "stderr": run_info["stderr"],
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["generated_completion"] = completion
                sample["full_code"] = full_code
                sample["raw_output"] = generated
            if self.save_ground_truth:
                sample["canonical_solution"] = canonical_solution
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "task_id": task_id,
                "prompt": prompt,
                "canonical_solution": canonical_solution,
                "generated_completion": completion,
                "passed": str(run_info["passed"]),
                "error": (run_info["stderr"] or run_info["error"] or "")[:500],
                "stdout": run_info["stdout"][:300],
            })

        pass_1 = sum(r["pass@1"] for r in results) / len(results)
        logger.info(f"HumanEval pass@1: {pass_1:.4f} ({len(results)} problems)")

        summary = {
            "benchmark": "humaneval",
            "pass@1": pass_1,
            "num_problems": len(results),
        }
        self._save_results(xlsx_rows, "humaneval", self.XLSX_COLUMNS, summary)

        return {**summary, "per_problem": results}

    def _extract_completion(self, generated: str, prompt: str) -> str:
        """Extract the function body completion from generated text."""
        if generated.startswith(prompt):
            return generated[len(prompt):]

        code_block = re.search(r"```(?:python)?\n(.*?)```", generated, re.DOTALL)
        if code_block:
            code = code_block.group(1)
            lines = code.split("\n")
            body_start = None
            for i, line in enumerate(lines):
                if line.strip().startswith("def "):
                    body_start = i + 1
                    break
            if body_start is not None:
                return "\n".join(lines[body_start:])
            return code

        if "def " in generated:
            lines = generated.split("\n")
            body_lines: list[str] = []
            in_body = False
            for line in lines:
                stripped = line.lstrip()
                if stripped.startswith("def "):
                    in_body = True
                    continue
                if in_body:
                    if stripped and not line[0].isspace() and not stripped.startswith("#"):
                        if stripped.startswith("def ") or stripped.startswith("class "):
                            break
                    body_lines.append(line)
            if body_lines:
                return "\n".join(body_lines)

        logger.debug(
            "HumanEval _extract_completion: no code fence / def found, "
            f"indenting raw output (first 120 chars): {generated[:120]!r}"
        )
        lines = generated.strip().split("\n")
        indented = []
        for line in lines:
            if line.strip():
                if not line.startswith(" ") and not line.startswith("\t"):
                    line = "    " + line
                indented.append(line)
            else:
                indented.append(line)
        return "\n".join(indented)

    def _run_humaneval_test(
        self, solution: str, test: str, entry_point: str, timeout: int = 10
    ) -> dict:
        """Run HumanEval test harness in a subprocess."""
        full_code = f"{solution}\n\n{test}\n\ncheck({entry_point})\n"
        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile(mode="w", suffix=".py", delete=False) as f:
                f.write(full_code)
                tmp_path = f.name
            result = subprocess.run(
                [sys.executable, tmp_path],
                capture_output=True, text=True, timeout=timeout,
            )
            return {
                "passed": result.returncode == 0,
                "timed_out": False,
                "stderr": result.stderr.strip(),
                "stdout": result.stdout.strip(),
                "error": "",
            }
        except subprocess.TimeoutExpired:
            return {"passed": False, "timed_out": True,
                    "stderr": "", "stdout": "", "error": f"timeout>{timeout}s"}
        except Exception as exc:
            return {"passed": False, "timed_out": False,
                    "stderr": "", "stdout": "", "error": str(exc)}
        finally:
            if tmp_path:
                Path(tmp_path).unlink(missing_ok=True)


# ──────────────────────────────────────────────────────────────────────────────
# HotpotQA
# ──────────────────────────────────────────────────────────────────────────────

class HotpotQAEvaluator(BenchmarkEvaluator):
    """HotpotQA: Multi-hop reasoning QA.

    Metrics: Exact Match (EM), F1
    Format: Given a question + context passages, answer the question.
    """

    SYSTEM_PROMPT = (
        "Answer the following question based on the given context. "
        "Think step by step, then provide a concise answer."
    )

    XLSX_COLUMNS = [
        "idx", "question", "context_summary",
        "ground_truth", "generated", "em", "f1",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        dataset = resolve_dataset("hotpot_qa", config="distractor", split="validation")
        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for idx, item in enumerate(tqdm(items, desc="HotpotQA")):
            question = item["question"]
            answer = item["answer"]
            context = item["context"]

            context_text = ""
            for title, sentences in zip(context["title"], context["sentences"]):
                context_text += f"[{title}]\n" + " ".join(sentences) + "\n\n"

            prompt = f"Context:\n{context_text}\nQuestion: {question}\nAnswer:"
            generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
            generated = generated_raw.strip().split("\n")[0]

            em = exact_match(generated, answer)
            f1 = f1_score(generated, answer)

            sample: dict[str, Any] = {
                "idx": idx,
                "question": question,
                "em": em,
                "f1": f1,
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["generated"] = generated
                sample["raw_output"] = generated_raw
            if self.save_ground_truth:
                sample["ground_truth"] = answer
                sample["context"] = context_text
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "idx": str(idx),
                "question": question,
                "context_summary": context_text[:300] + "…" if len(context_text) > 300 else context_text,
                "ground_truth": answer,
                "generated": generated,
                "em": str(em),
                "f1": f"{f1:.4f}",
            })

        avg_em = sum(r["em"] for r in results) / len(results)
        avg_f1 = sum(r["f1"] for r in results) / len(results)
        logger.info(f"HotpotQA EM: {avg_em:.4f}, F1: {avg_f1:.4f}")

        summary = {
            "benchmark": "hotpotqa",
            "exact_match": avg_em,
            "f1": avg_f1,
            "num_examples": len(results),
        }
        self._save_results(xlsx_rows, "hotpotqa", self.XLSX_COLUMNS, summary)

        return {**summary, "per_example": results}


# ──────────────────────────────────────────────────────────────────────────────
# MMLU
# ──────────────────────────────────────────────────────────────────────────────

class MMLUEvaluator(BenchmarkEvaluator):
    """MMLU: Massive Multitask Language Understanding.

    Metrics: Accuracy (5-shot)
    Format: Multiple choice questions (A/B/C/D).
    """

    SYSTEM_PROMPT = (
        "Answer the following multiple choice question. "
        "Think briefly, then respond with just the letter (A, B, C, or D)."
    )

    DEFAULT_SUBJECTS = [
        "abstract_algebra", "anatomy", "astronomy",
        "college_computer_science", "college_mathematics",
        "high_school_mathematics", "high_school_physics",
        "logical_fallacies", "machine_learning", "moral_scenarios",
    ]

    XLSX_COLUMNS = [
        "subject", "question", "choice_A", "choice_B", "choice_C", "choice_D",
        "ground_truth", "predicted", "correct",
    ]

    def __init__(self, *args, subjects: list[str] | None = None, **kwargs):
        super().__init__(*args, **kwargs)
        self.subjects = subjects or self.DEFAULT_SUBJECTS

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        all_correct = 0
        all_total = 0
        per_subject: dict[str, Any] = {}
        results: list[dict] = []
        xlsx_rows: list[dict] = []

        for subject in tqdm(self.subjects, desc="MMLU subjects"):
            try:
                dataset = resolve_dataset("cais/mmlu", config=subject, split="test")
                val_dataset = resolve_dataset("cais/mmlu", config=subject, split="validation")
            except Exception as e:
                logger.warning(f"Could not load MMLU subject '{subject}': {e}")
                continue

            items = list(dataset)
            if self.num_samples:
                items = items[:self.num_samples // len(self.subjects)]

            few_shot = self._build_few_shot(list(val_dataset)[:5])
            correct = 0

            for item in items:
                question = item["question"]
                choices = item["choices"]
                answer_idx = item["answer"]
                answer_letter = "ABCD"[answer_idx]

                prompt = self._format_mcq(question, choices, few_shot)
                generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
                pred = extract_multiple_choice(generated_raw)
                is_correct = pred == answer_letter
                if is_correct:
                    correct += 1

                sample: dict[str, Any] = {
                    "subject": subject,
                    "question": question,
                    "ground_truth": answer_letter,
                    "predicted": pred,
                    "correct": is_correct,
                }
                if self.save_qa:
                    sample["prompt"] = prompt
                    sample["raw_output"] = generated_raw
                if self.save_ground_truth:
                    sample["choices"] = {l: c for l, c in zip("ABCD", choices)}
                if self.record_trajectory:
                    sample["trajectory"] = trajectory

                results.append(sample)

                xlsx_rows.append({
                    "subject": subject,
                    "question": question,
                    "choice_A": choices[0] if len(choices) > 0 else "",
                    "choice_B": choices[1] if len(choices) > 1 else "",
                    "choice_C": choices[2] if len(choices) > 2 else "",
                    "choice_D": choices[3] if len(choices) > 3 else "",
                    "ground_truth": answer_letter,
                    "predicted": pred or "",
                    "correct": str(is_correct),
                })

            total = len(items)
            acc = correct / max(total, 1)
            per_subject[subject] = {"accuracy": acc, "n": total}
            all_correct += correct
            all_total += total

        overall_acc = all_correct / max(all_total, 1)
        logger.info(f"MMLU accuracy: {overall_acc:.4f} ({all_total} questions)")

        summary = {
            "benchmark": "mmlu",
            "accuracy": overall_acc,
            "num_questions": all_total,
            "per_subject": per_subject,
        }
        self._save_results(xlsx_rows, "mmlu", self.XLSX_COLUMNS, summary)

        return {**summary, "per_item": results}

    def _format_mcq(self, question: str, choices: list[str], few_shot: str = "") -> str:
        prompt = few_shot
        prompt += f"Question: {question}\n"
        for letter, choice in zip("ABCD", choices):
            prompt += f"{letter}. {choice}\n"
        prompt += "Answer:"
        return prompt

    def _build_few_shot(self, examples: list[dict]) -> str:
        text = ""
        for ex in examples:
            text += f"Question: {ex['question']}\n"
            for letter, choice in zip("ABCD", ex["choices"]):
                text += f"{letter}. {choice}\n"
            text += f"Answer: {'ABCD'[ex['answer']]}\n\n"
        return text


# ──────────────────────────────────────────────────────────────────────────────
# GSM8K
# ──────────────────────────────────────────────────────────────────────────────

class GSM8KEvaluator(BenchmarkEvaluator):
    """GSM8K: Grade School Math.

    Metrics: Accuracy (exact match on extracted number)
    Format: Solve a math word problem step by step.
    """

    SYSTEM_PROMPT = (
        "You are a math tutor. Solve the problem step by step, "
        "then give the final numerical answer after ####."
    )

    XLSX_COLUMNS = [
        "idx", "question", "ground_truth", "ground_truth_number",
        "generated", "extracted_number", "correct",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        dataset = resolve_dataset("openai/gsm8k", config="main", split="test")
        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for idx, item in enumerate(tqdm(items, desc="GSM8K")):
            question = item["question"]
            answer_raw = item["answer"]

            # Ground truth number is after "####"
            gt_number = answer_raw.split("####")[-1].strip().replace(",", "")

            prompt = (
                f"Solve the following math problem step by step.\n\n"
                f"Problem: {question}\n\nSolution:"
            )
            generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
            extracted = extract_number(generated_raw)

            is_correct = exact_match(extracted or "", gt_number)

            sample: dict[str, Any] = {
                "idx": idx,
                "correct": is_correct,
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["generated"] = generated_raw
                sample["extracted_number"] = extracted
            if self.save_ground_truth:
                sample["ground_truth"] = answer_raw
                sample["ground_truth_number"] = gt_number
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "idx": str(idx),
                "question": question,
                "ground_truth": answer_raw,
                "ground_truth_number": gt_number,
                "generated": generated_raw,
                "extracted_number": extracted or "",
                "correct": str(bool(is_correct)),
            })

        accuracy = sum(r["correct"] for r in results) / len(results)
        logger.info(f"GSM8K accuracy: {accuracy:.4f} ({len(results)} problems)")

        summary = {
            "benchmark": "gsm8k",
            "accuracy": accuracy,
            "num_problems": len(results),
        }
        self._save_results(xlsx_rows, "gsm8k", self.XLSX_COLUMNS, summary)

        return {**summary, "per_example": results}


# ──────────────────────────────────────────────────────────────────────────────
# MATH
# ──────────────────────────────────────────────────────────────────────────────

class MATHEvaluator(BenchmarkEvaluator):
    """MATH: Competition Mathematics.

    Metrics: Accuracy (exact match on extracted final answer)
    Format: Solve a competition math problem.
    The ground truth answer is inside \\boxed{...} in the solution field.
    """

    SYSTEM_PROMPT = (
        "You are a mathematics expert. Solve the problem and "
        "present your final answer inside \\boxed{}."
    )

    XLSX_COLUMNS = [
        "idx", "problem", "level", "type",
        "ground_truth", "generated", "correct",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        dataset = resolve_dataset("hendrycks/competition_math", split="test")
        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for idx, item in enumerate(tqdm(items, desc="MATH")):
            problem = item["problem"]
            solution = item["solution"]
            level = item.get("level", "")
            prob_type = item.get("type", "")

            # Extract ground truth from \boxed{...} in the solution
            gt_answer = self._extract_boxed(solution) or solution

            prompt = (
                f"Solve the following math problem.\n\n"
                f"Problem: {problem}\n\nSolution:"
            )
            generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)

            pred_answer = self._extract_boxed(generated_raw) or generated_raw
            is_correct = exact_match(pred_answer, gt_answer)

            sample: dict[str, Any] = {
                "idx": idx,
                "level": level,
                "type": prob_type,
                "correct": is_correct,
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["generated"] = generated_raw
            if self.save_ground_truth:
                sample["ground_truth"] = solution
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "idx": str(idx),
                "problem": problem,
                "level": level,
                "type": prob_type,
                "ground_truth": solution,
                "generated": generated_raw,
                "correct": str(bool(is_correct)),
            })

        accuracy = sum(r["correct"] for r in results) / len(results)
        logger.info(f"MATH accuracy: {accuracy:.4f} ({len(results)} problems)")

        summary = {
            "benchmark": "math",
            "accuracy": accuracy,
            "num_problems": len(results),
        }
        self._save_results(xlsx_rows, "math", self.XLSX_COLUMNS, summary)

        return {**summary, "per_example": results}

    @staticmethod
    def _extract_boxed(text: str) -> str | None:
        """Extract content from the last \\boxed{...} in the text."""
        # Find the last \boxed{...}, handling nested braces
        idx = text.rfind("\\boxed{")
        if idx == -1:
            return None
        start = idx + len("\\boxed{")
        depth = 1
        i = start
        while i < len(text) and depth > 0:
            if text[i] == "{":
                depth += 1
            elif text[i] == "}":
                depth -= 1
            i += 1
        if depth == 0:
            return text[start:i - 1].strip()
        return None


# ──────────────────────────────────────────────────────────────────────────────
# ARC
# ──────────────────────────────────────────────────────────────────────────────

class ARCEvaluator(BenchmarkEvaluator):
    """ARC: AI2 Reasoning Challenge (Challenge set).

    Metrics: Accuracy
    Format: Multiple choice science questions.
    """

    SYSTEM_PROMPT = (
        "Answer the following science question. "
        "Think briefly, then respond with just the letter."
    )

    XLSX_COLUMNS = [
        "idx", "question", "choices",
        "ground_truth", "predicted", "correct",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        dataset = resolve_dataset("allenai/ai2_arc", config="ARC-Challenge", split="test")
        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for idx, item in enumerate(tqdm(items, desc="ARC")):
            question = item["question"]
            labels = item["choices"]["label"]
            texts = item["choices"]["text"]
            answer_key = item["answerKey"]

            # Format MCQ
            prompt = f"Question: {question}\n"
            choices_str = ""
            for label, text in zip(labels, texts):
                prompt += f"{label}. {text}\n"
                choices_str += f"{label}. {text}; "
            prompt += "Answer:"

            generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
            pred = extract_multiple_choice(generated_raw)
            is_correct = pred == answer_key

            sample: dict[str, Any] = {
                "idx": idx,
                "question": question,
                "ground_truth": answer_key,
                "predicted": pred,
                "correct": is_correct,
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["raw_output"] = generated_raw
            if self.save_ground_truth:
                sample["choices"] = {l: t for l, t in zip(labels, texts)}
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "idx": str(idx),
                "question": question,
                "choices": choices_str.rstrip("; "),
                "ground_truth": answer_key,
                "predicted": pred or "",
                "correct": str(is_correct),
            })

        accuracy = sum(r["correct"] for r in results) / len(results)
        logger.info(f"ARC accuracy: {accuracy:.4f} ({len(results)} questions)")

        summary = {
            "benchmark": "arc",
            "accuracy": accuracy,
            "num_questions": len(results),
        }
        self._save_results(xlsx_rows, "arc", self.XLSX_COLUMNS, summary)

        return {**summary, "per_example": results}


# ──────────────────────────────────────────────────────────────────────────────
# ProntoQA
# ──────────────────────────────────────────────────────────────────────────────

class ProntoQAEvaluator(BenchmarkEvaluator):
    """ProntoQA: Propositional Logic QA.

    Metrics: Accuracy (exact match on True/False)
    Format: Determine if a conclusion follows from given premises.
    """

    SYSTEM_PROMPT = (
        "You are a logical reasoning expert. Determine if the conclusion "
        "follows from the given premises. Answer with True or False."
    )

    XLSX_COLUMNS = [
        "idx", "question", "ground_truth", "generated", "correct",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        try:
            dataset = resolve_dataset("renma/ProntoQA", split="test")
        except Exception:
            dataset = resolve_dataset("renma/ProntoQA", split="train")

        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for idx, item in enumerate(tqdm(items, desc="ProntoQA")):
            question = item.get("question", item.get("context", ""))
            answer = item.get("answer", item.get("label", ""))

            prompt = (
                f"{question}\n\n"
                f"Based on the above, is the conclusion true or false? Answer:"
            )
            generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
            generated = generated_raw.strip()

            is_correct = exact_match(generated, str(answer))

            sample: dict[str, Any] = {
                "idx": idx,
                "correct": is_correct,
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["generated"] = generated
                sample["raw_output"] = generated_raw
            if self.save_ground_truth:
                sample["ground_truth"] = str(answer)
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "idx": str(idx),
                "question": question,
                "ground_truth": str(answer),
                "generated": generated,
                "correct": str(bool(is_correct)),
            })

        accuracy = sum(r["correct"] for r in results) / len(results)
        logger.info(f"ProntoQA accuracy: {accuracy:.4f} ({len(results)} examples)")

        summary = {
            "benchmark": "prontoqa",
            "accuracy": accuracy,
            "num_examples": len(results),
        }
        self._save_results(xlsx_rows, "prontoqa", self.XLSX_COLUMNS, summary)

        return {**summary, "per_example": results}


# ──────────────────────────────────────────────────────────────────────────────
# GPQA
# ──────────────────────────────────────────────────────────────────────────────

class GPQAEvaluator(BenchmarkEvaluator):
    """GPQA: Graduate-level Google-Proof QA (Diamond subset).

    Metrics: Accuracy
    Format: Multiple choice (A/B/C/D), PhD-level science questions.
    """

    SYSTEM_PROMPT = (
        "Answer the following graduate-level science question. "
        "Think carefully, then respond with just the letter (A, B, C, or D)."
    )

    XLSX_COLUMNS = [
        "idx", "question", "choices",
        "ground_truth", "predicted", "correct",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        # GPQA Diamond is the hardest, expert-validated subset
        dataset = resolve_dataset("Idavidrein/gpqa", config="gpqa_diamond", split="train")
        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for idx, item in enumerate(tqdm(items, desc="GPQA")):
            question = item["Question"]
            correct_answer = item["Correct Answer"]
            choices = [
                item["Correct Answer"],
                item["Incorrect Answer 1"],
                item["Incorrect Answer 2"],
                item["Incorrect Answer 3"],
            ]
            # Deterministic shuffle based on question hash to keep reproducible
            import hashlib
            seed = int(hashlib.md5(question.encode()).hexdigest()[:8], 16)
            rng = __import__("random").Random(seed)
            rng.shuffle(choices)
            answer_key = "ABCD"[choices.index(correct_answer)]

            prompt = f"Question: {question}\n"
            choices_str = ""
            for letter, choice in zip("ABCD", choices):
                prompt += f"{letter}. {choice}\n"
                choices_str += f"{letter}. {choice[:60]}; "
            prompt += "Answer:"

            generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
            pred = extract_multiple_choice(generated_raw)
            is_correct = pred == answer_key

            sample: dict[str, Any] = {
                "idx": idx,
                "ground_truth": answer_key,
                "predicted": pred,
                "correct": is_correct,
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["raw_output"] = generated_raw
            if self.save_ground_truth:
                sample["choices"] = {l: c for l, c in zip("ABCD", choices)}
                sample["correct_answer_text"] = correct_answer
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "idx": str(idx),
                "question": question[:200],
                "choices": choices_str.rstrip("; "),
                "ground_truth": answer_key,
                "predicted": pred or "",
                "correct": str(is_correct),
            })

        accuracy = sum(r["correct"] for r in results) / max(len(results), 1)
        logger.info(f"GPQA accuracy: {accuracy:.4f} ({len(results)} questions)")

        summary = {
            "benchmark": "gpqa",
            "accuracy": accuracy,
            "num_questions": len(results),
        }
        self._save_results(xlsx_rows, "gpqa", self.XLSX_COLUMNS, summary)

        return {**summary, "per_example": results}


# ──────────────────────────────────────────────────────────────────────────────
# AIME (AMC/AIME Competition Math)
# ──────────────────────────────────────────────────────────────────────────────

class AIMEEvaluator(BenchmarkEvaluator):
    """AIME: American Invitational Mathematics Examination problems.

    Metrics: Accuracy (exact match on integer answer 000-999)
    Format: Competition math; answer is always an integer 0-999.
    """

    SYSTEM_PROMPT = (
        "You are a math competition expert. Solve the problem step by step. "
        "The answer is an integer between 000 and 999. "
        "Give your final answer as a 3-digit number after ####."
    )

    XLSX_COLUMNS = [
        "idx", "problem", "ground_truth", "extracted_answer", "correct",
    ]

    def evaluate(self) -> dict[str, Any]:
        from dllm_reason.utils.local_resolve import resolve_dataset

        # Use AI-MO/aimo-validation-aime which has AIME problems
        try:
            dataset = resolve_dataset("AI-MO/aimo-validation-aime", split="train")
        except Exception:
            # Fallback: use the broader competition math set
            try:
                dataset = resolve_dataset("Maxwell-Jia/AIME_2024", split="train")
            except Exception:
                logger.warning("Could not load AIME dataset — trying contest_math")
                dataset = resolve_dataset("hendrycks/competition_math", split="test")

        items = list(dataset)
        if self.num_samples:
            items = items[:self.num_samples]

        results = []
        xlsx_rows = []

        for idx, item in enumerate(tqdm(items, desc="AIME")):
            # Handle different dataset schemas
            problem = item.get("problem", item.get("question", ""))
            answer = str(item.get("answer", item.get("solution", "")))

            # For AIME, extract just the integer answer
            gt_number = self._extract_aime_answer(answer)

            prompt = (
                f"Solve the following competition math problem.\n\n"
                f"Problem: {problem}\n\n"
                f"Solution (give final integer answer after ####):"
            )
            generated_raw, trajectory = self._generate(prompt, self.SYSTEM_PROMPT)
            pred_number = extract_number(generated_raw) or ""

            is_correct = exact_match(str(pred_number), str(gt_number))

            sample: dict[str, Any] = {
                "idx": idx,
                "correct": is_correct,
            }
            if self.save_qa:
                sample["prompt"] = prompt
                sample["generated"] = generated_raw
                sample["extracted_answer"] = pred_number
            if self.save_ground_truth:
                sample["ground_truth"] = gt_number
                sample["full_solution"] = answer
            if self.record_trajectory:
                sample["trajectory"] = trajectory

            results.append(sample)

            xlsx_rows.append({
                "idx": str(idx),
                "problem": problem[:300],
                "ground_truth": str(gt_number),
                "extracted_answer": str(pred_number),
                "correct": str(bool(is_correct)),
            })

        accuracy = sum(r["correct"] for r in results) / max(len(results), 1)
        logger.info(f"AIME accuracy: {accuracy:.4f} ({len(results)} problems)")

        summary = {
            "benchmark": "aime",
            "accuracy": accuracy,
            "num_problems": len(results),
        }
        self._save_results(xlsx_rows, "aime", self.XLSX_COLUMNS, summary)

        return {**summary, "per_example": results}

    @staticmethod
    def _extract_aime_answer(text: str) -> str:
        """Extract integer answer from AIME solution."""
        # Try #### pattern first
        if "####" in text:
            return text.split("####")[-1].strip().replace(",", "")
        # Try \boxed{}
        idx = text.rfind("\\boxed{")
        if idx != -1:
            start = idx + len("\\boxed{")
            depth = 1
            i = start
            while i < len(text) and depth > 0:
                if text[i] == "{":
                    depth += 1
                elif text[i] == "}":
                    depth -= 1
                i += 1
            if depth == 0:
                return text[start:i - 1].strip()
        # Fallback: last number in text
        numbers = re.findall(r"(-?\d+)", text)
        return numbers[-1] if numbers else text.strip()


# ──────────────────────────────────────────────────────────────────────────────
# Registry
# ──────────────────────────────────────────────────────────────────────────────

BENCHMARK_REGISTRY = {
    "mbpp": MBPPEvaluator,
    "humaneval": HumanEvalEvaluator,
    "hotpotqa": HotpotQAEvaluator,
    "mmlu": MMLUEvaluator,
    "gsm8k": GSM8KEvaluator,
    "math": MATHEvaluator,
    "arc": ARCEvaluator,
    "prontoqa": ProntoQAEvaluator,
    "gpqa": GPQAEvaluator,
    "aime": AIMEEvaluator,
}
